import os
from typing import List

from celery.utils.log import get_task_logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from backend.enum import RequestStatus
from backend.models import (
    Comment,
    Complaint,
    Dossier,
    Municipality,
    News,
    Report,
    SubjectAccessRequest,
)
from etickets_v2.models import Agency
from polls.models import Poll
from settings.celery import app
from settings.cron_healthcheck import LoggingTask
from stats.functions import (
    days_since_last_instance,
    get_avg_final_response,
    get_avg_first_response,
    get_count_status,
    get_count_total,
    get_digital_tickets_count,
    get_officer_kpi_dashboard,
    get_percentage_closed_instances,
    get_percentage_closed_instances_in_less_than_specific_period,
    record_eticket_performance,
    record_operation_performance,
)

logger = get_task_logger(__name__)


@app.task(base=LoggingTask, name="operation_update_performance_table")
def operation_update_performance_table():
    """
    for each municipality
    generate operation update performance
    """
    municipalities = Municipality.objects.filter(is_active=True)
    for municipality in municipalities:
        try:
            record_operation_performance(
                municipality, [Complaint, Comment, SubjectAccessRequest]
            )
        except Exception as e:
            logger.error(f"operation_update_performance_table {municipality.id}, {e}")


@app.task(base=LoggingTask, name="etickets_performance_table")
def etickets_performance_table():
    agencies = Agency.objects.filter(is_active=True, municipality__is_active=True)
    for agency in agencies:
        try:
            record_eticket_performance(agency)
        except Exception as e:
            logger.error(f"etickets_performance_table {agencies.id}, {e}")


@app.task(base=LoggingTask, name="export_kpis_as_excel")
def export_kpis_as_excel(municipalities: List[dict], created_at, file_path):
    wb = Workbook()
    ws = wb.active

    categories = {
        "معطيات  عامة": [
            "ID",
            "اسم البلدية",
            "الولاية",
            "عدد المتساكنين",
            "تاريخ  تفعيل البلدية الرقمية",
            "عدد المسجلين بالمنصة",
            "عدد المتصرفين في المكتب الخلفي",
        ],
        "خدمة التبليغ عن مشكل": [
            "عدد مطالب التبليغ عن مشكل",
            "نسبة غلق مطالب التبليغ عن المشكل ",
            "نسبة غلق مطالب التبليغ في آجل 21 يوم",
            "معدل أيام التفاعل الأول ",
            "معدل أيام الإجابة النهائية",
            "قياس الآداء في علاقة بالخدمة ( عدد على 20)",
        ],
        "خدمة مطالب النفاذ إلى المعلومة": [
            "عدد مطالب النفاذ للمعلومة",
            "نسبة غلق مطالب النفاذ إلى المعلومة",
            "نسبة غلق مطالب النفاذ للمعلومة في آجل 20 يوم",
            "معدل أيام التفاعل الأول ",
            "معدل أيام الإجابة النهائية",
            "قياس الآداء في علاقة بالخدمة ( عدد على 20)",
        ],
        "خدمة الرخص": [
            "عدد ملفات الرخص",
            "عدد ملفات الرخص التي تم تقديمها و متابعتها",
            "نسبة غلق ملفات الرخص ",
            "نسبة غلق مطالب الرخص في آجل 30 يوم",
            "معدل أيام التفاعل الأول",
            "معدل أيام الإجابة النهائية",
            "قياس الآداء في علاقة بالخدمة ( عدد على 20)",
        ],
        "خدمة الأراء و التساؤلات": [
            "عدد الآراء و التساؤلات",
            "نسبة غلق الأراء و التساؤلات",
            "نسبة غلق الأراء و التساؤلات في أجال 30 يوم",
            "معدل أيام التفاعل الأول",
            "معدل أيام الإجابة النهائية",
            "قياس الآداء في علاقة بالخدمة ( عدد على 10)",
        ],
        "خدمة المستجدات": [
            "عدد المستجدات البلدية",
            "عدد الأيام التي لم يتم فيها تنزيل مستجدات منذ آخر منشور",
            "قياس الآداء في علاقة بالخدمة ( عدد على 10)",
        ],
        "خدمة تقارير المجلس البلدي ولجانه": [
            "عدد تقارير المجلس البلدي ولجانه",
            "عدد الأيام التي لم يتم فيها تنزيل تقارير المجلس البلدي ولجانه",
            "قياس الآداء في علاقة بالخدمة ( عدد على 10)",
        ],
        "خدمة مبادرات سبر الأراء": [
            "عدد مبادرات سبر الآراء",
            "عدد الأيام التي لم يتم فيها تنزيل مبادرات سبر أراء",
            "قياس الآداء في علاقة بالخدمة ( عدد على 10)",
        ],
        "معايير تقييم الأداء": [
            "عدد التذاكر الإلكترونية",
            "قياس الآداء العام للخدمات ( مجموع الآداء حسب الخدمات عدد 100 )",
            "معيار إنتاجية البلدية ( عدد  ملفات الرخص + عدد المستجدات + عدد التقارير + عدد مواعيد الأنشطة ) ( comparaison par rapport un seuil national ou autre )",
            "معيار إنتاجية المواطن ( عدد مطالب التبليغ عن المشكل + عدد مطالب النفاذ إلى المعلومة + عدد ملفات الرخص التي تمت متابعتها + عدد الآراء و التساؤلات + عدد مطالب الرخص + عدد التذاكر الإلكترونية)",
            "معيار الجودة ( نسبة غلق الملفات المعاجلة من رخص + مطالب تبليغ عن مشكل +مطالب نفاذ الى المعلومة ) آراء و تساؤلات a verifier ",
            "معيار الجودة ( معدل   أيام الإجابة الأولى ) ",
        ],
    }

    headers = []
    for category, kpis in categories.items():
        headers.append(category)
        headers.extend(kpis)
    # Add the headers to the worksheet
    row = 1
    col = 1
    for header in headers:
        if header in categories.keys():
            ws.merge_cells(
                start_row=row,
                start_column=col,
                end_row=row,
                end_column=col + len(categories[header]) - 1,
            )
            top_left_cell = ws.cell(row=row, column=col)
            top_left_cell.value = header
            top_left_cell.font = Font(bold=True)
        else:
            cell = ws.cell(row=row + 1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True)
        col += 1

    # Add data to the worksheet
    for municipality in municipalities:
        try:
            municipality_id = municipality['id']
            officer_kpi = get_officer_kpi_dashboard(municipality_id)
            complaints = Complaint.objects.filter(
                municipality_id=municipality_id, created_at__gt=created_at
            ).count()
            sars = SubjectAccessRequest.objects.filter(
                municipality_id=municipality_id, created_at__gt=created_at
            ).count()
            dossiers = Dossier.objects.filter(
                municipality_id=municipality_id, created_at__gt=created_at
            ).count()
            comments = Comment.objects.filter(
                municipality_id=municipality_id, created_at__gt=created_at
            ).count()
            news = News.objects.filter(
                municipality_id=municipality_id, published_at__gt=created_at
            ).count()
            reports = Report.objects.filter(
                municipality_id=municipality_id, created_at__gt=created_at
            ).count()
            polls = Poll.objects.filter(
                municipality_id=municipality_id, created_at__gt=created_at
            ).count()

            data = [
                # معطيات  عامة
                "",
                municipality_id,
                municipality['name'],
                municipality['city'],
                municipality['population'],
                municipality['activation_date'],
                municipality['total_followers'],
                municipality['total_managers'],
                # خدمة التبليغ عن مشكل
                "",
                complaints,
                get_percentage_closed_instances(officer_kpi["complaints"]),
                get_percentage_closed_instances_in_less_than_specific_period(
                    Complaint, officer_kpi["complaints"], municipality_id, 21
                ),
                get_avg_first_response(Complaint, municipality_id, created_at),
                get_avg_final_response(Complaint, municipality_id, created_at),
                "note /20",
                # خدمة مطالب النفاذ إلى المعلومة
                "",
                sars,
                get_percentage_closed_instances(officer_kpi["subject_access_requests"]),
                get_percentage_closed_instances_in_less_than_specific_period(
                    SubjectAccessRequest,
                    officer_kpi["subject_access_requests"],
                    municipality_id,
                    20,
                ),
                get_avg_first_response(
                    SubjectAccessRequest, municipality_id, created_at
                ),
                get_avg_final_response(
                    SubjectAccessRequest, municipality_id, created_at
                ),
                "note /20",
                # خدمة الرخص
                "",
                dossiers,
                get_count_status(
                    [RequestStatus.RECEIVED, RequestStatus.PROCESSING],
                    officer_kpi["dossiers"],
                ),
                get_percentage_closed_instances(officer_kpi["dossiers"]),
                get_percentage_closed_instances_in_less_than_specific_period(
                    Dossier, officer_kpi["dossiers"], municipality_id, 30
                ),
                get_avg_first_response(Dossier, municipality_id, created_at),
                get_avg_final_response(Dossier, municipality_id, created_at),
                "note /20",
                # خدمة الأراء و التساؤلات
                "",
                comments,
                get_percentage_closed_instances(officer_kpi["forum"]),
                get_percentage_closed_instances_in_less_than_specific_period(
                    Comment, officer_kpi["forum"], municipality_id, 30
                ),
                get_avg_first_response(Comment, municipality_id, created_at),
                get_avg_final_response(Comment, municipality_id, created_at),
                "note /10",
                # خدمة المستجدات
                "",
                news,
                days_since_last_instance(News, municipality_id),
                "note /10",
                # خدمة تقارير المجلس البلدي ولجانه
                reports,
                get_count_total(officer_kpi["reports"]),
                days_since_last_instance(Report, municipality_id),
                "note /10",
                # خدمة مبادرات سبر الأراء
                "",
                polls,
                days_since_last_instance(Poll, municipality_id),
                "note /10",
                # معايير تقييم الأداء
                "",
                get_digital_tickets_count(municipality_id),
                "note /100",
                "",
                "",
                "",
                "",
            ]
            ws.append(data)
        except Exception as e:
            logger.error(f"export_kpis_as_excel {municipality['id']}, {e}")
    try:
        wb.save(file_path)
    except Exception as e:
        logger.error(f"export_kpis_as_excel saving file, {e}")
    return file_path.replace("/backend", "")
